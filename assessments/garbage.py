from django.conf import settings
from django.db import models
import random
import os
from django.db.models.signals import pre_save, post_save
from django.dispatch import receiver
import openpyxl
import pandas as pd

from core.models import HashableModel


User = settings.AUTH_USER_MODEL
# Create your models here.



def get_filename_ext(filepath):
    base_name = os.path.basename(filepath)
    name, ext = os.path.splitext(base_name)
    return name, ext


def upload_file_path(instance, filename):
    # print(instance)
    # print(filename)
    new_filename = random.randint(1, 3910209312)
    name, ext = get_filename_ext(filename)
    final_filename = '{new_filename}{ext}'.format(
        new_filename=new_filename, ext=ext)
    return "test/{new_filename}/{final_filename}".format(
        new_filename=new_filename,
        final_filename=final_filename
    )
    
def cretificate_upload_file_path(instance, filename):
    # print(instance)
    # print(filename)
    new_filename = random.randint(1, 3910209312)
    name, ext = get_filename_ext(filename)
    final_filename = '{new_filename}{ext}'.format(
        new_filename=new_filename, ext=ext)
    return "cretificate/{new_filename}/{final_filename}".format(
        new_filename=new_filename,
        final_filename=final_filename
    )


class DriverCategory(HashableModel):
    name = models.CharField(max_length=20,null=True)
    en_passing_rate = models.PositiveIntegerField(default=0)
    pscy_passing_rate = models.PositiveIntegerField(default=0)

    class Meta:
        indexes = [
            models.Index(fields=['name','en_passing_rate','pscy_passing_rate']), 
        ]    
    def __str__(self):
        return self.name
    
TEST_TYPE = (
    ('en', 'en'),
    ('psyc', 'psyc'),
)

class Test(HashableModel):
    name = models.CharField(max_length=255)
    timestamp = models.DateTimeField(auto_now_add=True, null=True)
    type = models.CharField(max_length=10, choices=TEST_TYPE, default='')
    active = models.BooleanField(default=True)
    certificate_file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    english_Questions = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    number_question = models.PositiveIntegerField(default=1)
    maximum_attempts = models.PositiveIntegerField(default=5)
    psycometric_must = models.BooleanField(default=True)
    time_limit = models.PositiveIntegerField(default=7)
    second_chance = models.PositiveIntegerField(default=48)
    do_all_must = models.BooleanField(default=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['name','certificate_file', 'type','number_question','maximum_attempts','psycometric_must','do_all_must']), 
        ] 

    def __str__(self):
        return f'{self.name} {self.type}'

    def random_questions_mixed(self, number_of_questions, category=None):
        text_questions = self.question_set.filter(active=True, type='text')
        audio_questions = self.question_set.filter(active=True, type='audio')
        if category:
            text_questions = text_questions.filter(category=category)
            audio_questions = audio_questions.filter(category=category)
            
        number_of_text_questions = int(number_of_questions/2)
        number_of_audio_questions = number_of_questions - number_of_text_questions
        
        if number_of_text_questions > text_questions.count() or number_of_questions:
            number_of_text_questions = text_questions.count()
            # raise ValueError("Number of questions is greater than the number of questions available")
        if number_of_audio_questions > audio_questions.count() or number_of_questions :
            number_of_audio_questions = audio_questions.count()
        text_sample = random.sample(
            list(text_questions), number_of_text_questions)
        audio_sample = random.sample(
            list(audio_questions), number_of_audio_questions)
        return text_sample + audio_sample

    def random_questions(self, number_of_questions, question_type=None, category=None):
        questions = self.question_set.filter(active=True)
        if question_type:
            print(question_type)
            questions = questions.filter(type=question_type)
            print(questions)
        if category:
            questions = questions.filter(category=category)
            print(questions)
        # Check if number_of_questions is greater than the number of questions available
        if number_of_questions > questions.count() or number_of_questions < 0:
            number_of_questions = questions.count()
            # raise ValueError("Number of questions is greater than the number of questions available")
        return random.sample(list(questions), number_of_questions)


QUESTION_TYPE = (
    ('text', 'text'),
    ('audio', 'audio'),
)



@receiver(post_save, sender=Test)
def handle_uploaded_excel_file(sender,created, instance, **kwargs):
    # if created:
        print('postsave')
    # Check if the file was uploaded
        process_excel_file(instance.english_Questions.path, instance)
        
        if not instance.english_Questions:
            return

        # # Read the excel file into a pandas DataFrame
        df = pd.read_excel(instance.english_Questions.path)

        # Loop through each row of the DataFrame
        for index, row in df.iterrows():
            # Create a Question instance
            question = Question.objects.create(
                test=instance,
                text=row['Question Text'],
                type='MULTIPLE_CHOICE'
            )
            # Create Choice instances for each option
            for option in ['Option A', 'Option B', 'Option C', 'Option D']:
                is_correct = option == f"Option {row['Correct Answer']}"
                Choice.objects.create(
                    question=question,
                    text=row[option],
                    is_correct=is_correct
                )
post_save.connect(handle_uploaded_excel_file, sender=Test)


class Question(HashableModel):
    test = models.ForeignKey(Test, on_delete=models.CASCADE,null=True)
    active = models.BooleanField(default=True)
    text = models.TextField(null=True, blank=True)
    category = models.ManyToManyField(DriverCategory)
    type = models.CharField(max_length=10, choices=QUESTION_TYPE, default='')
    file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['text','test', 'type','file','active']), 
        ] 

    def __str__(self):
        return self.text


class Psycometric(HashableModel):
    test = models.ForeignKey(Test, on_delete=models.CASCADE)
    active = models.BooleanField(default=True)
    text = models.CharField(max_length=30)
    category = models.ManyToManyField(DriverCategory)
    file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    description = models.TextField(max_length=500, null=True, blank=True)
    timestamp = models.DateTimeField(auto_now_add=True, null=True)
    time_limit = models.PositiveIntegerField(default=7)
    must = models.BooleanField(default=True)

    class Meta:
        indexes = [
            models.Index(fields=['active','test','must', 'text','file','description','time_limit']), 
        ] 


    def __str__(self):
        return self.text
    


class Choice(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE,null=True)
    text = models.CharField(max_length=255,null=True)
    is_correct = models.BooleanField(default=False)

    def __str__(self):
        return self.text


class Answer(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE)
    choice = models.ForeignKey(Choice, on_delete=models.CASCADE)
    user = models.ForeignKey(User, on_delete=models.CASCADE)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        indexes = [
            models.Index(fields=['user','choice','created_at','question']), 
        ] 
        
    def __str__(self):
        return self.question.text

    def is_correct(self):
        return self.choice.is_correct

FINAL_MARK = (
    ('passed', 'passed'),
    ('falied', 'falied'),
    ('not_taken', 'not_taken'),
)

class TestAttempt(HashableModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, null=True, blank=True)
    test = models.ForeignKey(Test, on_delete=models.CASCADE, null=True, blank=True)
    psycometric = models.ForeignKey(Psycometric, on_delete=models.CASCADE, null=True, blank=True)
    type = models.CharField(max_length=10, choices=TEST_TYPE, default='')
    final_score = models.FloatField(default=0)
    timestamp = models.DateTimeField(auto_now_add=True,null=True)
    current_attempts = models.PositiveIntegerField(default=0)
    final_mark = models.CharField(max_length=10, choices=FINAL_MARK, default='not_taken')
    created_at= models.DateTimeField(auto_now=True,blank=True,null=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['user','test', 'type','final_score','timestamp','final_mark','created_at','psycometric']), 
        ] 
    
    def __str__(self):
        return f'{self.final_mark} {self.test} {self.psycometric}'        



class Cretificate(HashableModel):
    user = models.OneToOneField(User, on_delete=models.CASCADE, null=True, blank=True,related_name='certificate')
    test_attemp = models.ManyToManyField(TestAttempt)
    final_en_score = models.FloatField(default=0)
    final_psyco_score = models.FloatField(default=0)
    timestamp = models.DateTimeField(auto_now_add=True,null=True)
    created_at= models.DateTimeField(auto_now=True,blank=True,null=True)
    images =  models.ImageField(
        upload_to=upload_file_path, null=True, blank=True)
    

    class Meta:
        indexes = [
            models.Index(fields=['user', 'final_en_score','final_psyco_score','timestamp','created_at']), 
        ] 
    
    def __str__(self):
        return f'{self.user} => Psych: {self.final_psyco_score}, English: {self.final_en_score}'


    def check_certificate(self):
        if self.final_en_score >= self.user.driver_category.en_passing_rate and self.final_psyco_score >= self.user.driver_category.pscy_passing_rate:
            return True
        else:
            i = 0
            for marks in self.test_attemp:
                if marks.final_mark == 'passed':
                    i=+1
            if i == self.test_attemp.count():
                return True
            return False
        
        
        
        
@receiver(pre_save, sender=TestAttempt)
def increment_attempt(sender, instance, **kwargs):
    if instance.pk:
        return
    test = instance.test
    user = instance.user
    current_attempts = TestAttempt.objects.filter(user=user)
    current_attempts_count = current_attempts.count()
    user.current_attempts = current_attempts_count + 1
    if not user.is_superuser:
        user.test_active = False
    user.save()
    instance.current_attempt = current_attempts_count + 1
    


pre_save.connect(increment_attempt, sender=TestAttempt)





def process_excel_file(file_path,test):
    workbook = openpyxl.load_workbook(file_path)
    sheets = workbook.sheetnames

    for sheet in sheets:
        # Get the active sheet
        current_sheet = workbook[sheet]

        # Loop over the rows in the sheet
        for row in range(2, current_sheet.max_row + 1):
            # Get the category name from the sheet name
            category_name = sheet
            _ , driver_category = DriverCategory.objects.get_or_create(name=category_name)

            # Get the values from each cell in the row
            question_text = current_sheet.cell(row=row, column=2).value
            option_a = current_sheet.cell(row=row, column=3).value
            option_b = current_sheet.cell(row=row, column=4).value
            option_c = current_sheet.cell(row=row, column=5).value
            option_d = current_sheet.cell(row=row, column=6).value
            correct_answer = current_sheet.cell(row=row, column=7).value

            # Create a new Question instance
            question = Question.objects.create(
                test=test,
                active=True,
                text=question_text,
                type='text'
            )

            # Add the category to the Question instance
            question.category.add(driver_category)

            # Create the choices for the Question instance
            choice_a = Choice.objects.create(
                question=question,
                text=option_a,
                is_correct=(correct_answer == 'A')
            )
            choice_b = Choice.objects.create(
                question=question,
                text=option_b,
                is_correct=(correct_answer == 'B')
            )
            choice_c = Choice.objects.create(
                question=question,
                text=option_c,
                is_correct=(correct_answer == 'C')
            )
            choice_d = Choice.objects.create(
                question=question,
                text=option_d,
                is_correct=(correct_answer == 'D')
            )
            
            

def read_excel_file(file,test):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file)
    
    # Loop over each sheet in the workbook
    for sheet in workbook:
        # Get the category name from the sheet name
        category_name = sheet.title
        
        # Get or create the DriverCategory object
        category, created = DriverCategory.objects.get_or_create(name=category_name)
        
        # Loop over each row in the sheet (skip the first row with headers)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            # Read the question text, options, and correct answer from the row
            question_text = row[1].value
            option_a = row[2].value
            option_b = row[3].value
            option_c = row[4].value
            option_d = row[5].value
            correct_answer = row[6].value
            
            # Get or create the Question object
            question, created = Question.objects.get_or_create(test=test,text=question_text,type='text',actice=True)
            question.category.add(category)
            # Create the Choice objects for the options
            Choice.objects.get_or_create(question=question, text=option_a, defaults={'is_correct': correct_answer == 'A'})
            Choice.objects.get_or_create(question=question, text=option_b, defaults={'is_correct': correct_answer == 'B'})
            Choice.objects.get_or_create(question=question, text=option_c, defaults={'is_correct': correct_answer == 'C'})
            Choice.objects.get_or_create(question=question, text=option_d, defaults={'is_correct': correct_answer == 'D'})




































from django.conf import settings
from django.db import models
import random
import os
from django.db.models.signals import pre_save, post_save
from django.dispatch import receiver
import openpyxl
import pandas as pd

from core.models import HashableModel


User = settings.AUTH_USER_MODEL
# Create your models here.



def get_filename_ext(filepath):
    base_name = os.path.basename(filepath)
    name, ext = os.path.splitext(base_name)
    return name, ext


def upload_file_path(instance, filename):
    # print(instance)
    # print(filename)
    new_filename = random.randint(1, 3910209312)
    name, ext = get_filename_ext(filename)
    final_filename = '{new_filename}{ext}'.format(
        new_filename=new_filename, ext=ext)
    return "test/{new_filename}/{final_filename}".format(
        new_filename=new_filename,
        final_filename=final_filename
    )
    
def cretificate_upload_file_path(instance, filename):
    # print(instance)
    # print(filename)
    new_filename = random.randint(1, 3910209312)
    name, ext = get_filename_ext(filename)
    final_filename = '{new_filename}{ext}'.format(
        new_filename=new_filename, ext=ext)
    return "cretificate/{new_filename}/{final_filename}".format(
        new_filename=new_filename,
        final_filename=final_filename
    )


class DriverCategory(HashableModel):
    name = models.CharField(max_length=20,null=True)
    en_passing_rate = models.PositiveIntegerField(default=0)
    pscy_passing_rate = models.PositiveIntegerField(default=0)

    class Meta:
        indexes = [
            models.Index(fields=['name','en_passing_rate','pscy_passing_rate']), 
        ]    
    def __str__(self):
        return self.name
    
TEST_TYPE = (
    ('en', 'en'),
    ('psyc', 'psyc'),
)

class Test(HashableModel):
    name = models.CharField(max_length=255)
    timestamp = models.DateTimeField(auto_now_add=True, null=True)
    type = models.CharField(max_length=10, choices=TEST_TYPE, default='')
    active = models.BooleanField(default=True)
    certificate_file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    english_Questions = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    number_question = models.PositiveIntegerField(default=1)
    maximum_attempts = models.PositiveIntegerField(default=5)
    psycometric_must = models.BooleanField(default=True)
    time_limit = models.PositiveIntegerField(default=7)
    second_chance = models.PositiveIntegerField(default=48)
    do_all_must = models.BooleanField(default=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['name','certificate_file', 'type','number_question','maximum_attempts','psycometric_must','do_all_must']), 
        ] 

    def __str__(self):
        return f'{self.name} {self.type}'

    def random_questions_mixed(self, number_of_questions, category=None):
        text_questions = self.question_set.filter(active=True, type='text')
        audio_questions = self.question_set.filter(active=True, type='audio')
        if category:
            text_questions = text_questions.filter(category=category)
            audio_questions = audio_questions.filter(category=category)
            
        number_of_text_questions = int(number_of_questions/2)
        number_of_audio_questions = number_of_questions - number_of_text_questions
        
        if number_of_text_questions > text_questions.count() or number_of_questions:
            number_of_text_questions = text_questions.count()
            # raise ValueError("Number of questions is greater than the number of questions available")
        if number_of_audio_questions > audio_questions.count() or number_of_questions :
            number_of_audio_questions = audio_questions.count()
        text_sample = random.sample(
            list(text_questions), number_of_text_questions)
        audio_sample = random.sample(
            list(audio_questions), number_of_audio_questions)
        return text_sample + audio_sample

    def random_questions(self, number_of_questions, question_type=None, category=None):
        questions = self.question_set.filter(active=True)
        if question_type:
            print(question_type)
            questions = questions.filter(type=question_type)
            print(questions)
        if category:
            questions = questions.filter(category=category)
            print(questions)
        # Check if number_of_questions is greater than the number of questions available
        if number_of_questions > questions.count() or number_of_questions < 0:
            number_of_questions = questions.count()
            # raise ValueError("Number of questions is greater than the number of questions available")
        return random.sample(list(questions), number_of_questions)


QUESTION_TYPE = (
    ('text', 'text'),
    ('audio', 'audio'),
)


import pandas as pd




class QuestionCategory(HashableModel):
    text = models.TextField(null=True,blank=True)
    class Meta:
        indexes = [
            models.Index(fields=['text']), 
        ] 

    def __str__(self):
        return self.text

    
class Question(HashableModel):
    test = models.ForeignKey(Test, on_delete=models.CASCADE,null=True)
    active = models.BooleanField(default=True)
    text = models.TextField(null=True,blank=True)
    category = models.ManyToManyField(QuestionCategory)
    type = models.CharField(max_length=10, choices=QUESTION_TYPE, default='')
    file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['text','test', 'type','file','active']), 
        ] 

    def __str__(self):
        return self.text


    
class Psycometric(HashableModel):
    test = models.ForeignKey(Test, on_delete=models.CASCADE)
    active = models.BooleanField(default=True)
    text = models.CharField(max_length=30)
    category = models.ManyToManyField(DriverCategory)
    file = models.FileField(
        upload_to=upload_file_path, null=True, blank=True)
    description = models.TextField(max_length=500, null=True, blank=True)
    timestamp = models.DateTimeField(auto_now_add=True, null=True)
    time_limit = models.PositiveIntegerField(default=7)
    must = models.BooleanField(default=True)

    class Meta:
        indexes = [
            models.Index(fields=['active','test','must', 'text','file','description','time_limit']), 
        ] 


    def __str__(self):
        return self.text
    


class Choice(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE,null=True)
    text = models.CharField(max_length=255,null=True)
    is_correct = models.BooleanField(default=False)

    class Meta:
        indexes = [
            models.Index(fields=['text','question','is_correct']), 
        ] 
        
    def __str__(self):
        return self.text


class Answer(models.Model):
    question = models.ForeignKey(Question, on_delete=models.CASCADE)
    choice = models.ForeignKey(Choice, on_delete=models.CASCADE)
    user = models.ForeignKey(User, on_delete=models.CASCADE)
    created_at = models.DateTimeField(auto_now_add=True)

    class Meta:
        indexes = [
            models.Index(fields=['user','choice','created_at','question']), 
        ] 
        
    def __str__(self):
        return self.question.text

    def is_correct(self):
        return self.choice.is_correct

FINAL_MARK = (
    ('passed', 'passed'),
    ('falied', 'falied'),
    ('not_taken', 'not_taken'),
)

class TestAttempt(HashableModel):
    user = models.ForeignKey(User, on_delete=models.CASCADE, null=True, blank=True)
    test = models.ForeignKey(Test, on_delete=models.CASCADE, null=True, blank=True)
    psycometric = models.ForeignKey(Psycometric, on_delete=models.CASCADE, null=True, blank=True)
    type = models.CharField(max_length=10, choices=TEST_TYPE, default='')
    final_score = models.FloatField(default=0)
    timestamp = models.DateTimeField(auto_now_add=True,null=True)
    current_attempts = models.PositiveIntegerField(default=0)
    final_mark = models.CharField(max_length=10, choices=FINAL_MARK, default='not_taken')
    created_at= models.DateTimeField(auto_now=True,blank=True,null=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['user','test', 'type','final_score','timestamp','final_mark','created_at','psycometric']), 
        ] 
    
    def __str__(self):
        return f'{self.final_mark} {self.test} {self.psycometric}'        



class Cretificate(HashableModel):
    user = models.OneToOneField(User, on_delete=models.CASCADE, null=True, blank=True,related_name='certificate')
    test_attemp = models.ManyToManyField(TestAttempt)
    final_en_score = models.FloatField(default=0)
    final_psyco_score = models.FloatField(default=0)
    timestamp = models.DateTimeField(auto_now_add=True,null=True)
    created_at= models.DateTimeField(auto_now=True,blank=True,null=True)
    images =  models.ImageField(
        upload_to=upload_file_path, null=True, blank=True)
    

    class Meta:
        indexes = [
            models.Index(fields=['user', 'final_en_score','final_psyco_score','timestamp','created_at']), 
        ] 
    
    def __str__(self):
        return f'{self.user} => Psych: {self.final_psyco_score}, English: {self.final_en_score}'


    def check_certificate(self):
        if self.final_en_score >= self.user.driver_category.en_passing_rate and self.final_psyco_score >= self.user.driver_category.pscy_passing_rate:
            return True
        else:
            i = 0
            for marks in self.test_attemp:
                if marks.final_mark == 'passed':
                    i=+1
            if i == self.test_attemp.count():
                return True
            return False
        
        
        
        
@receiver(pre_save, sender=TestAttempt)
def increment_attempt(sender, instance, **kwargs):
    if instance.pk:
        return
    test = instance.test
    user = instance.user
    current_attempts = TestAttempt.objects.filter(user=user)
    current_attempts_count = current_attempts.count()
    user.current_attempts = current_attempts_count + 1
    if not user.is_superuser:
        user.test_active = False
    user.save()
    instance.current_attempt = current_attempts_count + 1
    


pre_save.connect(increment_attempt, sender=TestAttempt)





@receiver(post_save, sender=Test)
def handle_uploaded_excel_file(sender,created, instance, **kwargs):
    if not instance.type =="en" or not instance.english_Questions:
        Question.objects.all().delete()
        return  
    df = pd.read_excel(instance.english_Questions.path, sheet_name=None)

    # loop through each sheet
    # loop through each sheet
    for sheet_name, data in df.items():
        # process the data for the current sheet
        # extract the columns "Question Text", "Option A", "Option B", "Option C", "Option D", "Correct Answer"
        question_text = data["Question Text"]
        option_a = data["Option A"]
        option_b = data["Option B"]
        option_c = data["Option C"]
        option_d = data["Option D"]
        correct_answer = data["Correct Answer"]

        # create the Question and Choice objects and save them to the database
        questions = []
        choices = []
        for i in range(len(question_text)):
            # retrieve the category based on the sheet name
            category, created = QuestionCategory.objects.get_or_create(text=sheet_name)

            # create the Question object
            question = Question(
                test=instance,
                active=True,
                text=question_text[i],
                type="text",
                file=None
            )
            question.category.add(category)
            questions.append(question)

            # create the Choice objects
            choice_a = Choice(
                question=question,
                text=option_a[i],
                is_correct=correct_answer[i] == "A"
            )
            choice_b = Choice(
                question=question,
                text=option_b[i],
                is_correct=correct_answer[i] == "B"
            )
            choice_c = Choice(
                question=question,
                text=option_c[i],
                is_correct=correct_answer[i] == "C"
            )
            choice_d = Choice(
                question=question,
                text=option_d[i],
                is_correct=correct_answer[i] == "D"
            )
            choices.extend([choice_a, choice_b, choice_c, choice_d])

        Question.objects.bulk_create(questions)
        Choice.objects.bulk_create(choices)

            
post_save.connect(handle_uploaded_excel_file, sender=Test)























import mimetypes
import xlwt
def audio_qustion(request):    
    if request.method == 'POST':
        test_id = request.POST.get('test_id')
        audio_file = request.FILES.get('audioFile')
        json_file = request.FILES.get('json')
        questions_file = request.FILES.get('Questions')
        with io.StringIO(audio_file.read().decode('utf-8')) as file:
            audio_file = json.load(file)
            # print(audio_file)
            audio_data = []
            for data in audio_file:
                id = data["id"]
                name = data["name"]
                audio_data.append([id, name])
        question_records = []

        questions_data = json.loads(questions_file.read().decode('utf-8'))
        for question in questions_data:
            name = question["name"]
            options = re.findall(r's:\d+:\"([^\"]+)\";', question["options"])
            correct_answer = int(re.findall(r'\d+', question["answer"])[0])
            text = question.get("text")
            file_number = re.search(r'\b\d{4}\b', text)
            if file_number:
                number = file_number.group()
                for file_name in audio_data:
                    if file_name[0] == number:
                        file_path = file_name[1]
            else:
                number = 0000
            question_records.append([name, options, options[correct_answer], file_path, number])  
        
        with open('json.xlsx', 'wb') as f:
            wb = xlwt.Workbook(encoding='utf-8')
            ws = wb.add_sheet('Audio Questions')
            row_num = 0
            columns = ['Name', 'Option A', 'Option B', 'Option C', 'Option D', 'Correct Answer', 'Audio File Path', 'File Number']
            for col_num, column_title in enumerate(columns):
                ws.write(row_num, col_num, column_title)
            for row in question_records:
                row_num += 1
                ws.write(row_num, 0, row[0])
                ws.write(row_num, 1, row[1][0])
                ws.write(row_num, 2, row[1][1])
                ws.write(row_num, 3, row[1][2])
                ws.write(row_num, 4, row[1][3])
                ws.write(row_num, 5, row[2])
                ws.write(row_num, 6, row[3])
                ws.write(row_num, 7, row[4])
            wb.save(f)
        
        with open('json.xlsx', 'rb') as f:
            file_data = f.read()
            
        instance = Test.objects.get(id=test_id)
        instance.audio_Questions.save('json.xlsx', ContentFile(file_data))
        instance.save()
        return HttpResponse("File uploaded successfully")

    # return JsonResponse({'success': False})
    return render(request, 'assessments/audio.html')
