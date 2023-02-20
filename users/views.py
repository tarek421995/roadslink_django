from django.conf import settings
from django.contrib.auth import (
    authenticate,
    get_user_model,
    login,
    logout,
)
from django.http import HttpResponseRedirect, JsonResponse, HttpResponse

from django.contrib.auth.hashers import make_password
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
from django.utils.crypto import get_random_string
from django.shortcuts import (
    get_object_or_404,
    redirect,
    render,
)
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

from django.db.models import Q

from assessments.models import DriverCategory, TestAttempt
from django.db.models import Prefetch
from .forms import (
    CustomLoginForm,
    DriversRegisterForm,
    EditUserForm,
    RegisterForm,
    ForgetPasswordEmailCodeForm,
    ChangePasswordForm,
    OtpForm,
)
from .models import AgentCategory, CompanyCategory, DriverAge, DriverNationality, Invocies, InvociesFile, OtpCode, CustomUser,Cretificate
from .utils import (
    send_activation_code,
    send_reset_password_code,
)
from .decorators import only_authenticated_user, redirect_authenticated_user


User = get_user_model()

@only_authenticated_user
def home_view(request):
    return render(request, 'shared_layout/home.html')


@redirect_authenticated_user
def login_view(request):
    error = None
    c = {'next' : request.GET.get('next', '/')}
    print(c)
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request, username=form.cleaned_data['username_or_email'], password=form.cleaned_data['password'])
            if user:
                if not user.is_active:
                    messages.warning(request, _(
                        f"It's look like you haven't still verify your email - {user.email}"))
                    return redirect('users:activate_email')
                else:
                    login(request, user)
                    if 'next' in request.POST:
                        return redirect(request.POST['next'])
                    else:
                        print('redirect here')
                        # return HttpResponseRedirect('/')
                        return redirect('users:home')
            else:
                error = 'Invalid Credentials'
    else:
        form = CustomLoginForm()
    return render(request, 'users/login.html', {'form': form, 'error': error})


@only_authenticated_user
@login_required
def logout_view(request):
    logout(request)
    return redirect('users:login')


@redirect_authenticated_user
def registeration_view(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST or None)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.source = 'Register'
            user.save(True)

            code = get_random_string(20)
            otp = OtpCode(code=code, user=user)
            otp.save(True)
            try:
                send_activation_code(user.email, code)
            except:
                otp.delete()
                user.delete()
                messages.error(request, _('Failed while sending code!'))
            else:
                messages.success(
                    request, _(f'We have sent a verification code to your email - {user.email}'))
                return redirect('users:activate_email')
    else:
        form = RegisterForm()
    return render(request, 'users/register.html', {'form': form})

def edit_driver_username(request, id):
    print(id)
    user = get_user_model().objects.get(id=id)
    # user = get_user_model().objects.get(username=username)
    if request.method == 'POST':
        form = DriversRegisterForm(request.POST,instance=user)
        if form.is_valid():
            
            form.save()
            messages.success(request,'you have successfully edit the current user')
            return redirect('users:driver_registeration_view')
        else:
            messages.warning(request,'something in the  data formate wrong')

        
    else:
        form = DriversRegisterForm(instance=user)    
    return render(request, 'users/register_staff.html', {'form': form, 'custom': True})

def register(request):
    print('register')
    if request.method == 'POST':
        print('register')
        username = request.POST.get('username')
        
        driver_category = request.POST.get('driver_category')
        contact = request.POST.get('contact')
        company = request.POST.get('company')
        # agent = request.POST.get('agent')
        age = request.POST.get('age')
        nationality = request.POST.get('nationality')
        full_name = request.POST.get('full_name')
        offical_ID_number = request.POST.get('offical_ID_number')
        driver_category = DriverCategory.objects.get(name=driver_category)
        # Cat_agent = AgentCategory.objects.get(name=agent)

        # company = get_object_or_404(CompanyCategory,name=company)
        comapny = CompanyCategory.objects.filter(name=company)
        if company:
            company = comapny.first()
        else:
            messages.warning(request,'you have to fill the fields before submting')
            test = 'you have to fill the fields before submting'
            return JsonResponse({'status':test})

        age = DriverAge.objects.get(number=age)
        nationality = DriverNationality.objects.get(name=nationality)
        user = User.objects.filter(
        Q(username__contains=username) | 
        Q (offical_ID_number__contains=offical_ID_number)
        )
        print(user.count())
        # user1 = 
        
        if user.count()>=1:
            messages.warning(request,'the user is already exist')
            return JsonResponse({'status': 'redirect' ,'id':user.first().id})
        else:
            user = User.objects.create(username=username,driver_category=driver_category,offical_ID_number=offical_ID_number,contact=contact,company=company, age=age,nationality=nationality)
            user.set_password('123456')
        user.test_active = True
        user.email = f'{username}_'
        user.full_name = full_name
        user.is_active = True
        user.source = 'Driver_Registeration_Roadslink'
        user.save()
        # TO_DO
        
        # process the data and save to database
        print(username)            
        return JsonResponse({'status': 'success'})
    
    driver_categories= DriverCategory.objects.all()   
    companies  = CompanyCategory.objects.all()
    driver_ages  = DriverAge.objects.all()
    driver_nationality  = DriverNationality.objects.all()
    return render(request, 'users/driver_registr.html', {'driver_categories': driver_categories,'driver_ages':driver_ages,'driver_nationality':driver_nationality, 'companies': companies})
  
  
  
def edit_profile(request):
    error = None
    if request.method == 'POST':
        form = EditUserForm(request.POST,request.FILES, instance=request.user)
        print(form.data)
        if form.is_valid():
            user = request.user
            user.profile_picture = form.data['profile_picture']
            form.save()
            print(user.profile_picture)
            return redirect('users:edit_profile')
        else:
            error = 'Invalid data'
    else:
        form = EditUserForm(instance=request.user)
        
    return render(request, 'users/user_edit.html', {'form': form, 'error':error})
  
@redirect_authenticated_user
def forgot_password_view(request):
    if request.method == 'POST':
        form = ForgetPasswordEmailCodeForm(request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data['username_or_email']
            user = get_user_model().objects.get(**username_or_email)
            code = get_random_string(20)

            otp = OtpCode(code=code, user=user, email=user.email)
            otp.save()

            try:
                send_reset_password_code(user.email, code)
            except:
                otp.delete()
                messages.error(request, _('Failed while sending code!'))
            else:
                messages.success(request, _(
                    f"We've sent a passwrod reset otp to your email - {user.email}"))
                return redirect('users:reset_code')
    else:
        form = ForgetPasswordEmailCodeForm()
    return render(request, 'users/forgot_password.html', context={'form': form})

@redirect_authenticated_user
def check_otp_view(request):
    if request.method == 'POST':
        form = OtpForm(request.POST)
        if form.is_valid():
            otp = OtpCode.objects.get(code=form.cleaned_data['otp'])
            user = otp.user
            otp.delete()
            user.is_active = True
            user.save()
            return redirect('users:login')
    else:
        form = OtpForm()
    return render(request, 'users/user_otp.html', {'form': form})

@redirect_authenticated_user
def check_reset_otp_view(request):
    if request.method == 'POST':
        form = OtpForm(request.POST)
        if form.is_valid():
            otp = OtpCode.objects.get(code=form.cleaned_data['otp'])
            request.session['email'] = otp.user.email
            messages.success(request, _(
                "Please create a new password that you don't use on any other site."))
            return redirect('users:reset_new_password')
    else:
        form = OtpForm()
    return render(request, 'users/user_otp.html', {'form': form})


@redirect_authenticated_user
def reset_new_password_view(request):
    if request.method == 'POST':
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            email = request.session['email']
            del request.session['email']
            user = CustomUser.objects.get(email=email)
            user.password = make_password(form.cleaned_data["new_password2"])
            user.save()
            messages.success(request, _(
                "Your password changed. Now you can login with your new password."))
            return redirect('users:login')
    else:
        form = ChangePasswordForm()
    return render(request, 'users/new_password.html', {'form': form})



def certificate_print(request):
    users = User.objects.all().order_by('-last_login').prefetch_related(
        Prefetch('certificate', queryset=Cretificate.objects.only('final_en_score', 'final_psyco_score')))
    if request.user.is_staff:
        return render(request, 'users/printing.html', {'users': users})
    return render(request, 'users/printing.html')

# def check_user_certificate(user):
#     if hasattr(user, 'certificate'):
#         return True
#     else:
#         return False

def user_search(request):
    search_value = request.GET.get('search', '')
    user_data = User.objects.filter(
        Q(username__contains=search_value) | 
        Q(email__contains=search_value) | 
        Q(full_name__contains=search_value) | 
        Q(nationality__name__contains=search_value) | 
        Q(company__name__contains=search_value)) \
                        .values('id', 'username', 'company__name', 'test_active', 'last_login', 'current_attempts','nationality__name', 'is_active') \
                        .prefetch_related('company')
    user_ids = [user['id'] for user in user_data]
    certificate_data = Cretificate.objects.filter(user_id__in=user_ids) \
                        .values('user_id', 'final_en_score', 'final_psyco_score')

    certificate_data_dict = {cert['user_id']: cert for cert in certificate_data}
    
    data = [{
        'id': user['id'],
        'username': user['username'],
        'company': user['company__name'],
        'last_login': user['last_login'],
        'current_attempts': user['current_attempts'],
        'test_active': user['test_active'],
        'has_certificate': bool(certificate_data_dict.get(user['id'], {}).get('final_en_score', '') and certificate_data_dict.get(user['id'], {}).get('final_psyco_score', '')),
        'final_en_score': certificate_data_dict.get(user['id'], {}).get('final_en_score', ''),
        'final_psyco_score': certificate_data_dict.get(user['id'], {}).get('final_psyco_score', ''),
        'is_active': user['is_active'],
    } for user in user_data]
    print(data)
    return JsonResponse(data, safe=False)



@login_required
def get_companies(request):
    input = request.GET.get('input', '')
    companies = CompanyCategory.objects.filter(
        Q(name__contains=input)
        ).values() 
    companies_list = list(companies)
    print(companies_list)
    return JsonResponse(companies_list, safe=False)

@login_required
def get_agents(request):
    company_name = request.GET.get('company_name')
    print(company_name)
    agents = AgentCategory.objects.filter(companycategory__name=company_name).values('id', 'name')
    print(agents)
    data = {'agents': list(agents)}
    return JsonResponse(data, safe=False)
    # return JsonResponse(agents_list, safe=False)

import openpyxl
    
from django.utils import timezone
from datetime import datetime ,date
@login_required
def generate_user_invoice(request,user_id,company_id):
    user = User.objects.get(id=user_id)
    company=CompanyCategory.objects.get(id=company_id)
    create_by = request.user
    try:
        invocies_file = InvociesFile.objects.get(active=True,type='indvidual')
    except:
        messages.warning(request, 'you have to add invoice file first')
        return redirect('users:certificate_print')
    
    invoice,created  = Invocies.objects.get_or_create(create_by=create_by.full_name,type='indvidual' ,company=company,user=user)
    
    if created:
        invoice.issued_date = timezone.now()
    invoice.save()
    template = openpyxl.load_workbook(invocies_file.file)
    sheet = template.active
    today = date.today().strftime("%mm/%d/%Y")
    # Fill in the user data
    sheet['D11'] = 1
    sheet['B7'] = user.full_name
    sheet['B8'] = company.name
    sheet['E11'] = invocies_file.fees
    sheet['F4'] = f'54551{invoice.id}'
    sheet['F3'] = today    
    # Generate the response with the filled Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=invoice_Roadslink_{invoice.id}.xlsx'
    template.save(response)
    return response


@login_required
def generate_company_invoice(request):
    invoices = Invocies.objects.filter(active=True)
    companies = CompanyCategory.objects.all()
    context = {'list_invoice':invoices,'company':companies}
    if request.method == 'POST':
        try:
            users = User.objects.all()
            invocies_file = InvociesFile.objects.get(active=True,type='company')
            template = openpyxl.load_workbook(invocies_file.file)
            sheet = template.active
            new_invoice = invoices.filter(create_by=request.user,type='company')
            start_date = request.POST.get('start_date', '')
            end_date = request.POST.get('end_date', '')
            company_name = request.POST.get('company_name', '')
            if not start_date and not end_date and not company_name:
                messages.warning(request, 'Please specify the start and end date and company name.')
                return render(request, 'users/generate_invoices.html', context)
            if start_date and end_date:
                start_date_parts = start_date.split('-')
                end_date_parts = end_date.split('-')
                start_datetime = datetime(int(start_date_parts[0]), int(start_date_parts[1]), int(start_date_parts[2]), 0, 0, 0)
                end_datetime = datetime(int(end_date_parts[0]), int(end_date_parts[1]), int(end_date_parts[2]), 23, 59, 59)
                start_datetime = timezone.make_aware(start_datetime, timezone.get_default_timezone())
                end_datetime = timezone.make_aware(end_datetime, timezone.get_default_timezone())
                invoices = invoices.filter(timestamp__range=[start_datetime, end_datetime])
                users = users.filter(date_joined__range=[start_datetime, end_datetime])
                print('old_user',users)
                print('invoices',invoices) 
                
                if not company_name:
                    invoices = invoices.filter(type='group')
                    old_users = set()
                    for invoice in invoices:
                        old_users.update(invoice.users.all())
                            
                    if len(old_users) > 0 :
                    # print('old_users len',len(old_users))
                        users = users.exclude(id__in=[u.id for u in old_users])
                        
                    if len(users) > 0 :
                        type= 'group'
                        internal_invoices = Invocies.objects.create(type=type,start_date=start_datetime,end_date=end_datetime)
                        internal_invoices.create_by=request.user.username
                        for user in users:
                            internal_invoices.users.add(user)
                        internal_invoices.save()
                    else:
                        messages.warning(request, f'you have old invoice invoice generated  {start_date} to {end_date}, note : this invoice is not for B2B offical email')
                        return render(request, 'users/generate_invoices.html', context)
                    
                    sheet['C9'] = f'from {start_date} to {end_date}'
                    today = date.today().strftime("%m/%d/%Y")
                    sheet['D11'] = users.count()
                    sheet['B7'] = company_name
                    sheet['E11'] = invocies_file.fees
                    sheet['F4'] = f'3345{internal_invoices.id}'
                    sheet['F3'] = today
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename=invoice_Roadslink _{company_name}_{users.count()}.xlsx'
                    template.save(response)
                    messages.success(request, f'you have successfully create invoice for internal use case from {start_date} to {end_date}, note : this invoice is not for B2B offical email')
                    print('invoices', invoices)   
                    return response   
                            
            invoices = invoices.filter(type='indvidual')
            if company_name:
                invoices = invoices.filter(company__name=company_name)
                users = users.filter(company__name=company_name)
                new_invoice = new_invoice.filter(company__name=company_name)
                
            excluded_users = invoices.values_list('user__id', flat=True)            
            users = users.exclude(id__in=excluded_users)
            if company_name and new_invoice.count() > 0:
                old_users = set()
                for invoice in new_invoice:
                    if invoice.type == 'company' and invoice.company.name == company_name:
                        old_users.update(invoice.users.all())
                if len(old_users) > 0 :
                    users = users.exclude(id__in=[u.id for u in old_users])
                    if users.count() <= 0 :
                        messages.warning(request, f'There are no more users to make invoices for this company or timeline ')
                        return render(request, 'users/generate_invoices.html', context)
                    else:
                        company = get_object_or_404(CompanyCategory, name=company_name)
                        type= 'company'
                        sheet['B7'] = company_name
                        new_invoice = Invocies.objects.create(create_by=request.user.username,type=type,company=company)
                        for user in users:
                            new_invoice.users.add(user)
                        new_invoice.save()  
                else:
                    messages.warning(request, f'There are no more users to make invoices for this company or timeline ')
            else:
                if users.count() <= 0 :
                    messages.warning(request, f'There are no more users to make invoices for this company or timeline ')
                else:
                    if company_name:
                        company = get_object_or_404(CompanyCategory, name=company_name)
                        type= 'company'
                        sheet['B7'] = company_name
                        new_invoice = Invocies.objects.create(create_by=request.user.username,type=type,company=company)
                    for user in users:
                        new_invoice.users.add(user)
                    new_invoice.save()                
            today = date.today().strftime("%m/%d/%Y")
            sheet['D11'] = users.count()
            sheet['E11'] = invocies_file.fees
            sheet['F4'] = f'3345{new_invoice.id}'
            sheet['F3'] = today
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=invoice_Roadslink _{company_name}_{today}.xlsx'
            template.save(response)
            messages.success(request,'Invoices Generated Successfully')
            print('invoices', invoices)   
            return response             
    
        except ValueError as ex:
            messages.warning(request, f'Invalid input: {str(ex)}')
            return render(request, 'users/generate_invoices.html', context)
        
        except Exception as ex:
            print(str(ex))
            messages.error(request, f'Unexpected error: {str(ex)}')
            return render(request, 'users/generate_invoices.html', context)
        
    return render(request, 'users/generate_invoices.html', context)
            
@login_required
def downalod_invoice(request,id):
    invoice = Invocies.objects.get(id=id)
    template = openpyxl.load_workbook(invoice.invocies_file.file)
    sheet = template.active
    
    # Fill in the user data
    sheet['D11'] = invoice.user.count()
    if invoice.invocies_file.type == 'company' and invoice.company:
        sheet['B7'] = invoice.company.name
        sheet['B6'] = invoice.company.agent_set().values('name')
        sheet['C9'] = f'from {invoice.start_date} to {invoice.end_date}'
        sheet['F4'] = f'3354{invoice.id}'

    elif invoice.invocies_file.type == 'group':
        sheet['C9'] = f'from {invoice.start_date} to {invoice.end_date}'
        sheet['F4'] = f'3345{invoice.id}'

    else:
        sheet['B7'] = invoice.user.first()
        sheet['F4'] = f'54551{invoice.id}' 
    today = date.today().strftime("%m/%d/%Y")
    sheet['F3'] = today   
    
    sheet['B8'] = invoice.invocies_file.location
    sheet['E11'] = invoice.invocies_file.fees

    # Generate the response with the filled Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=invoice_Roadslink_{invoice.id}.xlsx'
    template.save(response)
    return response
    
    